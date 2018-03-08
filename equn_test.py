from openpyxl import Workbook
import datetime
import time
import ahrs
import globalv
            
def send_PWM(ser,PWM1,PWM2):
	try:
		ardfl = 0
        	to_send = str(PWM1) + ":" + str(PWM2) + ":|"
        	ser.write(to_send.encode())
        	ser.flushInput()
	except:
		ardfl = 1
		pass

def ang_acc():
	g = ahrs.gyro()
	vel = g[2]
	time.sleep(0.01)
	g = ahrs.gyro()
	acc = abs(abs(g[2]) - abs(vel))/0.01
	return acc

def data_test(stage):
    if (stage == 0)
        tflag = 0
        tinc = 1
        pwm = 1500
        d = 0
        t = time()
		log ['A'+str(i)] = "Stage 1 log start"
		i = i+1
		log ['A'+str(i)] = "date/time"
		log ['B'+str(i)] = "Diff. PWM"
		log ['C'+str(i)] = "ang vel"
		i = i+1
        while (tflag = 1):
            if d > 400:
                break
            d = d+30
            while (abs(time () - t) < tinc):   
                send_PWM(serthr , (pwm +d) , (pwm-d))
				wb = Workbook()
				log = wb.active()
				log ['A'+str(i)] = datetime.datetime.now()
				log ['B'+str(i)] = d
				g = ahrs.gyro()
				log ['C'+str(i)] = g[2] 
				log ['D'+str(i)] = ang_acc()
				i = i+1

	if (stage == 1)
		tflag = 0
        tinc = 1
        pwm = 1500
        d = 0
        t = time()
		log ['A'+str(i)] = "Stage 2 log start"
		i = i+1
		log ['A'+str(i)] = "date/time"
		log ['B'+str(i)] = "Diff. PWM"
		log ['C'+str(i)] = "ang vel"
		log ['D'+str(i)] = "ang acc"
		i = i+1
        while (tflag = 1):
            if d > 400:
                break
            d = d+30
            while (abs(time () - t) < tinc):   
                send_PWM(serthr , (pwm -d) , (pwm+d))
				wb = Workbook()
				log = wb.active()
				log ['A'+str(i)] = datetime.datetime.now()
				log ['B'+str(i)] = d				
				g = ahrs.gyro()
				log ['C'+str(i)] = g[2]
				log ['D'+str(i)] = ang_acc()
				i = i+1

	if (stage == 2)
		tflag = 0
        tinc = 20
        pwm = 1500
        d = 400
        t = time()
		log ['A'+str(i)] = "Stage 3 log start"
		i = i+1
		log ['A'+str(i)] = "date/time"
		log ['B'+str(i)] = "Diff. PWM"
		log ['C'+str(i)] = "ang vel"
		i = i+1
		while (abs(time () - t) < tinc):   
			send_PWM(serthr , (pwm -d) , (pwm+d))
			wb = Workbook()
			log = wb.active()
			log ['A'+str(i)] = datetime.datetime.now()
			log ['B'+str(i)] = d				
			g = ahrs.gyro()
			log ['C'+str(i)] = g[2]
			log ['D'+str(i)] = ang_acc()
			i = i+1

		d = 0
		tinc = 30
		while (abs(time () - t) < tinc):   
			send_PWM(serthr , (pwm -d) , (pwm+d))
			wb = Workbook()
			log = wb.active()
			log ['A'+str(i)] = datetime.datetime.now()
			log ['B'+str(i)] = d				
			g = ahrs.gyro()
			log ['C'+str(i)] = g[2]
			log ['D'+str(i)] = ang_acc()
			i = i+1
	
	if (stage == 3)
		tflag = 0
        tinc = 20
        pwm = 1500
        d = 400
        t = time()
		log ['A'+str(i)] = "Stage 4 log start"
		i = i+1
		log ['A'+str(i)] = "date/time"
		log ['B'+str(i)] = "Diff. PWM"
		log ['C'+str(i)] = "ang vel"
		i = i+1
		while (abs(time () - t) < tinc):   
			send_PWM(serthr , (pwm -d) , (pwm+d))
			wb = Workbook()
			log = wb.active()
			log ['A'+str(i)] = datetime.datetime.now()
			log ['B'+str(i)] = d				
			g = ahrs.gyro()
			log ['C'+str(i)] = g[2]
			log ['D'+str(i)] = ang_acc()
			i = i+1

		d = 0
		tinc = 30
		while (abs(time () - t) < tinc):   
			send_PWM(serthr , (pwm -d) , (pwm+d))
			wb = Workbook()
			log = wb.active()
			log ['A'+str(i)] = datetime.datetime.now()
			log ['B'+str(i)] = d				
			g = ahrs.gyro()
			log ['C'+str(i)] = g[2]
			log ['D'+str(i)] = ang_acc()
			i = i+1

	wb.save(str(datetime.datetime.now())+"RUN log")

